"""
文件核对工具 v2
功能：从 Excel/CSV 读取文件名清单，扫描目标文件夹（含子文件夹），智能匹配并报告缺失文件。

用法:
  # 交互模式（不带参数，按提示输入）
  python check_files.py

  # 命令行模式（最简）
  python check_files.py 表格.xlsx 目标文件夹

  # 自动检测"文件名"所在列（扫描表头找匹配关键词）
  python check_files.py 表格.xlsx 目标文件夹 --auto

  # 手动指定列、起始行、工作表
  python check_files.py 表格.xlsx 目标文件夹 -c 2 -r 2 -s "修订计划"

  # 指定报告输出路径（默认输出到当前目录）
  python check_files.py 表格.xlsx 目标文件夹 -o 报告.txt

  # 同时输出 CSV 格式报告（方便后续处理）
  python check_files.py 表格.xlsx 目标文件夹 --csv 结果.csv
"""

import argparse
import csv
import sys
from pathlib import Path

# 自动检测列时匹配的关键词（按优先级排列）
HEADER_KEYWORDS = ["文件名", "文件", "文档名", "名称", "文档", "标题"]


# ========================
# 📖 表格读取
# ========================

def detect_filename_column(header_row):
    """扫描表头，自动找到文件名所在列。返回 1-based 列号，找不到返回 None。"""
    for keyword in HEADER_KEYWORDS:
        for i, cell in enumerate(header_row):
            if cell and keyword in str(cell).strip():
                return i + 1
    return None


def read_filenames_from_excel(filepath, column=None, start_row=None, sheet_name=None):
    """从 Excel 读取文件名列表。支持自动检测列和工作表选择。"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 缺少 openpyxl，请运行: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # 选择工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"❌ 找不到工作表 '{sheet_name}'，可用工作表: {', '.join(wb.sheetnames)}")
            wb.close()
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        if len(wb.sheetnames) > 1:
            print(f"💡 当前工作表: '{ws.title}'（共 {len(wb.sheetnames)} 个: {', '.join(wb.sheetnames)}）")
            print(f"   如需切换，加参数 -s 工作表名")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # 自动检测列：扫描前几行找表头
    if column is None:
        for header_row in rows[:3]:  # 最多看前3行
            detected = detect_filename_column(header_row)
            if detected:
                column = detected
                # 如果检测到表头，数据从下一行开始
                if start_row is None:
                    start_row = detect_header_row(rows, detected)
                print(f"🔍 自动检测到「文件名」在第 {column} 列（{chr(64+column) if column <= 26 else '列'+str(column)}）")
                break
        if column is None:
            print("⚠️  未能自动检测文件名列，默认使用第 2 列（B列）。如不对请用 -c 指定")
            column = 2

    if start_row is None:
        start_row = 2  # 默认跳过表头

    filenames = []
    for row in rows[start_row - 1:]:
        if not row:
            continue
        idx = column - 1
        if idx >= len(row):
            continue
        name = row[idx]
        if name and str(name).strip():
            clean = str(name).strip().replace('\n', ' ')
            # 去重
            if clean not in filenames:
                filenames.append(clean)
    return filenames


def detect_header_row(rows, column):
    """推断数据从第几行开始（找表头所在行）。"""
    idx = column - 1
    for i, row in enumerate(rows[:5]):
        if idx < len(row) and row[idx] and any(kw in str(row[idx]) for kw in HEADER_KEYWORDS):
            return i + 2  # 表头在 i 行（0-based），数据从 i+2 行（1-based）开始
    return 2


def read_filenames_from_csv(filepath, column=None, start_row=None):
    """从 CSV 读取文件名列表。自动尝试多种编码。"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030", "big5"]
    rows = None
    used_encoding = None

    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                reader = list(csv.reader(f))
            rows = reader
            used_encoding = enc
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if rows is None:
        print(f"❌ 无法解码 CSV 文件，尝试过编码: {', '.join(encodings)}")
        sys.exit(1)

    if len(rows) <= 1:
        return []

    # 自动检测列
    if column is None:
        detected = detect_filename_column(rows[0])
        if detected:
            column = detected
            if start_row is None:
                start_row = 2
            print(f"🔍 自动检测到「文件名」在第 {column} 列")
        else:
            print("⚠️  未能自动检测文件名列，默认使用第 2 列。如不对请用 -c 指定")
            column = 2

    if start_row is None:
        start_row = 2

    filenames = []
    idx = column - 1
    for row in rows[start_row - 1:]:
        if len(row) > idx and row[idx].strip():
            clean = row[idx].strip().replace('\n', ' ')
            if clean not in filenames:
                filenames.append(clean)
    return filenames


# ========================
# 🔍 文件扫描与匹配
# ========================

def get_all_files_in_folder(folder):
    """递归获取文件夹中所有文件。"""
    folder = Path(folder)
    return [f for f in folder.rglob("*") if f.is_file()]


def find_matching_files(target_name, all_files):
    """
    智能查找匹配的文件：
    1. 精确匹配（忽略大小写）
    2. 前缀模糊匹配（处理 name → name_20260101、name-最终版 等）
    """
    target_path = Path(target_name)
    target_stem = target_path.stem.lower()
    target_suffix = target_path.suffix.lower()
    has_ext = bool(target_suffix)

    matches = []
    for f in all_files:
        f_stem = f.stem.lower()
        f_name = f.name.lower()

        # 1. 精确匹配
        if has_ext and f_name == target_name.lower():
            matches.append(f)
            continue
        if not has_ext and f_stem == target_stem:
            matches.append(f)
            continue

        # 2. 前缀模糊匹配（仅当表格没写后缀时触发）
        if not has_ext and f_stem.startswith(target_stem) and len(f_stem) > len(target_stem):
            next_char = f_stem[len(target_stem)]
            # 紧跟着的字符必须是非字母（允许 _、-、空格、数字等分隔符）
            if not next_char.isalpha():
                matches.append(f)
                continue

    return matches


# ========================
# 📊 输出
# ========================

def format_size(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"


def print_report(found_list, missing_list, target_folder, total_scanned):
    """控制台输出核对结果。"""
    print(f"\n{'='*65}")
    print(f"  📊 核对结果汇总")
    print(f"{'='*65}")
    print(f"  ✅ 找到: {len(found_list)} 个")
    print(f"  ❌ 缺失: {len(missing_list)} 个")
    print(f"  📁 扫描文件总数: {total_scanned}")

    if found_list:
        print(f"\n{'─'*65}")
        print(f"✅ 已找到的文件:")
        print(f"{'─'*65}")
        for item in found_list:
            exts = item['exts']
            matches = item['matches']

            if len(exts) > 1:
                ext_tip = f"⚠️ 多种格式: {', '.join(exts)}"
            else:
                ext_tip = f"格式: {exts[0]}"

            print(f"  🟢 {item['name']}")
            print(f"     [{ext_tip}] (共 {len(matches)} 个)")
            for f in matches:
                try:
                    rel_path = f.relative_to(target_folder)
                except ValueError:
                    rel_path = f
                size = format_size(f.stat().st_size)
                print(f"     📄 {rel_path}  ({size})")

    if missing_list:
        print(f"\n{'─'*65}")
        print(f"❌ 缺失的文件（文件夹及子文件夹中均未找到）:")
        print(f"{'─'*65}")
        for name in missing_list:
            print(f"  🔴 {name}")


def save_txt_report(filepath, found_list, missing_list, target_folder, total_target, total_scanned):
    """保存 TXT 格式报告。"""
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("📂 文件核对报告\n")
        f.write(f"表格目标: {total_target} 个 | 找到: {len(found_list)} | 缺失: {len(missing_list)} | 扫描: {total_scanned}\n")
        f.write("=" * 60 + "\n\n")

        if missing_list:
            f.write("【❌ 缺失文件】\n")
            for name in missing_list:
                f.write(f"  - {name}\n")
            f.write("\n")

        if found_list:
            f.write("【✅ 已找到文件】\n")
            for item in found_list:
                f.write(f"  [{item['name']}] 格式: {', '.join(item['exts'])}\n")
                for m in item['matches']:
                    try:
                        rel = m.relative_to(target_folder)
                    except ValueError:
                        rel = m
                    f.write(f"    -> {rel}\n")
            f.write("\n")


def save_csv_report(filepath, found_list, missing_list, target_folder):
    """保存 CSV 格式报告（方便后续数据处理）。"""
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["表格文件名", "状态", "匹配文件路径", "格式", "文件大小"])
        for item in found_list:
            if item['matches']:
                for m in item['matches']:
                    try:
                        rel = m.relative_to(target_folder)
                    except ValueError:
                        rel = m
                    ext = m.suffix.lower() if m.suffix else "(无后缀)"
                    writer.writerow([
                        item['name'], "找到", str(rel), ext, format_size(m.stat().st_size)
                    ])
            else:
                writer.writerow([item['name'], "找到", "", "", ""])
        for name in missing_list:
            writer.writerow([name, "缺失", "", "", ""])


# ========================
# 🚀 主流程
# ========================

def interactive_prompt(args):
    """交互模式：没有传参数时引导用户输入。"""
    print("=" * 55)
    print("  📂 文件核对工具 v2（交互模式）")
    print("=" * 55)

    if not args.table:
        path = input("📄 表格文件路径（.xlsx / .csv）: ").strip().strip('"').strip("'")
        if not path:
            print("未输入路径，退出。")
            sys.exit(0)
        args.table = path

    if not args.folder:
        path = input("📁 目标文件夹路径: ").strip().strip('"').strip("'")
        if not path:
            print("未输入路径，退出。")
            sys.exit(0)
        args.folder = path

    if not args.auto and not args.column:
        choice = input("🔍 自动检测文件名列？(回车=自动 / n=手动输入列号): ").strip().lower()
        if choice in ("n", "no", "否"):
            col = input("   文件名在第几列？(A=1, B=2...): ").strip()
            try:
                args.column = int(col)
            except ValueError:
                print("列号无效，将使用自动检测")
                args.auto = True
        else:
            args.auto = True

    if args.auto:
        args.column = None  # None 触发自动检测


def main():
    parser = argparse.ArgumentParser(
        description="文件核对工具：从表格读取文件名清单，扫描文件夹，智能匹配并报告缺失文件。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s 表格.xlsx 目标文件夹           # 最简用法
  %(prog)s 表格.xlsx 目标文件夹 --auto    # 自动检测列
  %(prog)s 表格.xlsx 目标文件夹 -c 2 -r 2 # 指定列和起始行
  %(prog)s 表格.xlsx 目标文件夹 -s 工作表  # 指定Excel工作表
  %(prog)s 表格.xlsx 目标文件夹 -o 报告.txt
  %(prog)s 表格.xlsx 目标文件夹 --csv 结果.csv
  %(prog)s                                  # 交互模式
        """,
    )
    parser.add_argument("table", nargs="?", help="表格文件路径（.xlsx / .csv）")
    parser.add_argument("folder", nargs="?", help="目标文件夹路径")
    parser.add_argument("-c", "--column", type=int, help="文件名所在列（A=1, B=2...），不填则自动检测")
    parser.add_argument("-r", "--start-row", type=int, help="数据起始行（跳过表头），默认 2")
    parser.add_argument("-s", "--sheet", help="Excel 工作表名称（不填用第一个）")
    parser.add_argument("-o", "--output", help="报告输出路径（默认: 当前目录/文件核对报告.txt）")
    parser.add_argument("--csv", dest="csv_output", help="同时输出 CSV 格式报告")
    parser.add_argument("--auto", action="store_true", help="自动检测文件名所在列（扫描表头关键词）")
    args = parser.parse_args()

    # 交互模式
    if not args.table:
        interactive_prompt(args)

    table_file = Path(args.table).resolve()
    target_folder = Path(args.folder).resolve()

    # 切换工作目录到脚本所在位置（方便相对路径报告输出）
    if args.output:
        report_path = Path(args.output).resolve()
    else:
        report_path = Path.cwd() / "文件核对报告.txt"

    # 校验
    print(f"\n{'='*65}")
    print(f"  📂 文件核对工具 v2")
    print(f"{'='*65}")
    if not table_file.exists():
        print(f"❌ 找不到表格文件: {table_file}")
        return
    if not target_folder.exists():
        print(f"❌ 找不到目标文件夹: {target_folder}")
        return

    print(f"📄 表格文件: {table_file}")
    print(f"📁 目标文件夹: {target_folder} (含所有子文件夹)")
    if args.sheet:
        print(f"📋 工作表: {args.sheet}")
    print("-" * 65)

    # 读取表格
    print("\n⏳ 正在读取表格并扫描文件夹...")
    suffix = table_file.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        if suffix == ".xls":
            print("⚠️  .xls 旧格式支持有限，建议另存为 .xlsx。尝试读取中...")
        column = args.column if not args.auto else None
        filenames = read_filenames_from_excel(
            table_file, column=column, start_row=args.start_row, sheet_name=args.sheet
        )
    elif suffix == ".csv":
        column = args.column if not args.auto else None
        filenames = read_filenames_from_csv(
            table_file, column=column, start_row=args.start_row
        )
    else:
        print(f"❌ 不支持的表格格式: {suffix}（仅支持 .xlsx / .csv）")
        return

    all_files = get_all_files_in_folder(target_folder)
    print(f"✅ 读取到 {len(filenames)} 个目标文件，扫描到 {len(all_files)} 个实际文件")
    print("-" * 65)

    # 核对
    found_list = []
    missing_list = []
    for name in filenames:
        matches = find_matching_files(name, all_files)
        if not matches:
            missing_list.append(name)
        else:
            exts = sorted(set(m.suffix.lower() if m.suffix else "(无后缀)" for m in matches))
            found_list.append({"name": name, "matches": matches, "exts": exts})

    # 输出
    print_report(found_list, missing_list, target_folder, len(all_files))

    # 保存报告
    save_txt_report(report_path, found_list, missing_list, target_folder, len(filenames), len(all_files))
    print(f"\n💾 报告已保存: {report_path}")

    if args.csv_output:
        csv_path = Path(args.csv_output).resolve()
        save_csv_report(csv_path, found_list, missing_list, target_folder)
        print(f"💾 CSV 报告: {csv_path}")

    print("=" * 65)


if __name__ == "__main__":
    main()
